from __future__ import annotations

import io

from shared.types import ExtractionError, ExtractionResult


def extract_xlsx(data: bytes) -> ExtractionResult:
    """Extract text from XLSX using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"## {sheet_name}\n" + "\n".join(rows))

    wb.close()

    if not sheets:
        raise ExtractionError("empty_content", "XLSX has no data")

    full_text = "\n\n".join(sheets)
    return ExtractionResult(
        text=full_text,
        metadata={
            "sheet_count": len(sheets),
            "word_count": len(full_text.split()),
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
